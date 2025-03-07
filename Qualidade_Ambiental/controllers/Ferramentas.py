
if 0==1:  # Este namespace serve apenas para a IDE enchergar e trabalhar com os itens abaixo
    from gluon import *
    from gluon import db, SQLFORM, IS_IN_SET, IS_UPPER, IS_EMPTY_OR, IS_IN_DB, IS_NOT_IN_DB, IS_MATCH, redirect, URL, FORM_CPF
    from gluon import XML, a_db, db, auth, Auth, Field, buscador, analise_de_residuos_projedatos, geradocspy, scrap_aprova_dig_
    request = current.request
    response = current.response
    session = current.session
    cache = current.cache
    T = current.T

from os import remove, path
from re import sub, match
from datetime import datetime
from gluon.contrib.pymysql.err import IntegrityError
from pathlib import Path
from openpyxl import load_workbook # type: ignore
from scrap_aprova_dig_ import raspa_aprova # type: ignore
from folder_maker import criar_pasta_compartilhada # type: ignore
from gluon import SQLFORM

mensagem_contrução = "Em Contrução! 🛠"

def processar_arquivo_html(arquivo):
    try:
        dados_do_processo = raspa_aprova('file:///{}/{}'.format(str(path.join(request.folder, 'uploads')), arquivo))
        print('Dados obtidos apartir do arquivo!')
        return dados_do_processo
    except Exception as e:
        session.flash = f'Erro ao processar o arquivo enviado: {e}'
        return None


def remover_arquivo(arquivo):
    try:
        remove('{}/{}'.format(str(path.join(request.folder, 'uploads')), arquivo))
        print('Arquivo removido da memória.')
    except OSError as e:
        session.flash = f"Erro na remoção do arquivo do server: Error: {e.strerror} : \
        {path.join(request.folder, 'uploads')} : {arquivo}"


def atualizar_planilha_control(dados_do_processo):
    planilha_control = Path(pasta_qualidade, 'DOCUMENTOS', 'Controle de Processos.xlsx') # type: ignore
    print('Planilha encontrada')
    wb = load_workbook(planilha_control, rich_text=True)
    print('Planilha aberta')
    ws = wb.active
    print('Planilha ativada')
    hj = datetime.today()
    ws.append([hj.strftime('%d/%m/%Y') , dados_do_processo['protocolo'], dados_do_processo['dados_gerador']['Nome'], 1,
                1 if float(dados_do_processo['dados_obra']['area_construida_final']) > 400.00 else 0, 0,
                        dados_do_processo['dados_obra']['area_construida_final']])
    try:
        print('Dados inseridos')
        wb.save(planilha_control)
    except PermissionError as pe:
        print('Planilha bloqueada')
        response.flash = 'Planilha de controle aberta. Feche e repita o procedimento.'
        redirect(URL(c='static', f='423.html' ))
    print('Planilha de processos atualizada.')


def buscar_ou_inserir_pessoa(dados_gerador):
    cnpj = dados_gerador.get('CNPJ')
    cpf = dados_gerador.get('CPF') or dados_gerador.get('cpf')
    cpf = cpf.replace('.','').replace('-','') if cpf else None
    nome = dados_gerador.get('Nome') or dados_gerador.get('nome')
    telefone = dados_gerador.get('tel') or ''
    email = dados_gerador.get('email') or ''

    if cnpj:
        pessoa = db.executesql(f"SELECT Id, Nome FROM Pessoas WHERE Pessoas.CNPJ = '{cnpj}' LIMIT 1;") or \
        db.executesql(f"SELECT Id, Nome FROM Pessoas WHERE Pessoas.CNPJ = '{cnpj.replace('.','').replace('/','').replace('-','')}' LIMIT 1;") or \
        db.executesql(f"SELECT Id, Nome FROM Pessoas WHERE Pessoas.CNPJ = '{'{}{}.{}{}{}.{}{}{}/{}{}{}{}-{}{}'.format(*cnpj)}' LIMIT 1;")
    elif cpf:
        pessoa = db.executesql(f"SELECT Id, Nome FROM Pessoas WHERE Pessoas.CPF = '{cpf}' LIMIT 1;") or \
        db.executesql(f"SELECT Id, Nome FROM Pessoas WHERE Pessoas.CPF = '{cpf.replace('.','').replace('-','')}' LIMIT 1;")
    else:
        pessoa = None

    if pessoa:
        print(f'Pessoa ::{pessoa[0][1]}:: encontrada no banco de dados!')
        return pessoa[0][0]
    else:

        pessoa = db.Pessoas.validate_and_insert(Nome=nome, CNPJ=cnpj,
                                                CPF= cpf, Telefone=telefone ,
                                                Email=email)
        db.commit()

        print(f'Pessoa :::{pessoa.id}: inserida no banco de dados!')
        return pessoa.id

def buscar_ou_inserir_processo(protocolo, id_pessoa, depto, serv):
    processo = db.executesql(f"SELECT Id FROM Processos WHERE Processos.protocolo = '{protocolo}' LIMIT 1 ;")
    if processo:
        print(f'Processo ::{protocolo}:: encontrado no banco de dados!')
        return processo[0][0]
    else:
        processo = db.Processos.validate_and_insert(Protocolo=protocolo,
                                                    IdPessoa=id_pessoa,
                                                    IdDpto=depto,
                                                    IdTipo=serv)
        db.commit()
        print(f'Processo ::{protocolo}:: inserido no banco de dados!')
        return processo.id

def buscar_ou_inserir_logradouro(endereco_obra):
    cep = endereco_obra.get('cep') or endereco_obra.get('Cep') or endereco_obra.get('CEP')
    try:
        cep = int(sub( '[^0-9]','',cep))
    except Exception as e:
        print(f'Problema com CEP:: {e}')
    logradouro = db.executesql(f"SELECT Id, Logradouro FROM Logradouros WHERE Logradouros.Cep = '{cep}' LIMIT 1 ;")
    if logradouro:
        print(f'Logradouro ::{logradouro[0][1]}:: encontrado no banco de dados!')
        return logradouro[0][0]
    else:
        bairro = endereco_obra.get('bairro') or endereco_obra.get('Bairro') or endereco_obra.get('BAIRRO') or ''
        nomebairro = sub(' - .*$', '' , bairro)
        palavra_chave_bairro = ' '.join(nomebairro.split()[ int(len(nomebairro.split())//2): int(len(nomebairro.split())/2+2) ])
        try:
            if len(palavra_chave_bairro) > 1:
                idbairro = db.executesql(f"SELECT Bairros.Id, Bairros.Bairro FROM Bairros WHERE Bairros.Bairro LIKE '%{palavra_chave_bairro}%';", as_dict= True)
            else:
                idbairro = {}
            if len(idbairro) > 0:
                idbairro, nomebairro = idbairro[0].get('Id'),  idbairro[0].get('Bairro')
                print(f'Bairro :: {nomebairro}:: encontrado no banco de dados com id [{idbairro}]!')
            else:
                print(f"Bairro {bairro} não encontrado")
                bairro_id = db.Bairros.validate_and_insert(
                    Bairro = bairro,
                    Cor = '', IdCidade= 9999
                    )

        except Exception as e:
                    session.flash = "Erro ao cadastrar Bairro"
                    print(e)
                    print('Redirecionando para form de Logradouros')
                    redirect(URL('default', 'Logradouros'))

        logradouro = endereco_obra.get('Logradouro') or endereco_obra.get('LOGRADOURO')
        elementos_de_logradouro = logradouro.split()
        if elementos_de_logradouro[0].upper().startswith('R'):
            denomin = 'RUA'
        elif elementos_de_logradouro[0].upper().startswith('AV'):
            denomin = 'AVENIDA'
        else:
            denomin = '-'
        try:
            logradouro = db.Logradouros.validate_and_insert(
                            Logradouro = ' '.join(logradouro.split()[1:] if denomin != '-' else logradouro),
                            Cep = cep,
                            Denominacao = denomin,
                            Prefixo='-',
                            IdBairro= idbairro if idbairro else bairro_id,
                            IdCidade=9999)
            print(f'Logradouro ::{logradouro}:: inserido no banco de dados!')

            db.commit()
            return logradouro.id
        except Exception as e:
                    session.flash = f"Erro ao cadastrar Logradouro: {e}"
                    print(f'erro ao inserir {logradouro} -> {e}')
                    redirect(URL('default', 'Logradouros'))



def buscar_ou_inserir_endereco(id_logradouro, endereco_obra):
    quadra = endereco_obra.get('Quadra') or endereco_obra.get('QUADRA') or ''
    lote = endereco_obra.get('Lote') or endereco_obra.get('LOTE') or ''
    num = endereco_obra.get('NUM') or endereco_obra.get('Num') or 0
    num = int(sub( '[^0-9]','',num))
    idendereco =    db.Enderecos((db.Enderecos.IdLogradouro == id_logradouro) & (db.Enderecos.Num == num)) \
                        or db.Enderecos((db.Enderecos.IdLogradouro == id_logradouro) &
                        (db.Enderecos.Num == num) & (db.Enderecos.Quadra == quadra ) & (db.Enderecos.Lote == lote ))\
                        or  db.Enderecos((db.Enderecos.IdLogradouro == id_logradouro) &
                        (db.Enderecos.Quadra == '0' + quadra) & (db.Enderecos.Lote == '0' + lote ) &
                        (db.Enderecos.Num == num))
    if idendereco:
        print(f'Endereço ::{idendereco.Id}:: encontrado no banco de dados!')
        return idendereco.id
    elif db.Enderecos((db.Enderecos.IdLogradouro == id_logradouro) &
                        (db.Enderecos.Num == num)) and num!= 0:
        idendereco = db.Enderecos((db.Enderecos.IdLogradouro == id_logradouro) & (db.Enderecos.Num == num ))
        print(f'Endereço ::{idendereco.Id}:: encontrado no banco de dados!')
        return idendereco.id
    else:
        try:
            idendereco = db.Enderecos.validate_and_insert(IdLogradouro = id_logradouro,
                Num = num, Quadra = quadra, Lote = lote, Complemento = '')
            db.commit()
            print(f'Endereço ::{idendereco}:: inserido no banco de dados!')
            return idendereco.id
        except IntegrityError as ie:
            session.flash = f'Tentou registrar o Endereço mas este Endereço já está Registrado. Erro: {ie.args}'

        except Exception as e:
            session.flash = f'{e}'
            redirect(URL('default', 'Enderecos'))


def traduz_data(data):
    import locale
    locale.setlocale(locale.LC_ALL, 'pt_BR.utf8')
    padroes_data = {'a': '%d/%m/%Y', 'b': '%d-%m-%Y', 'c': r'%d\%m\%Y',
                   'd': '%d/%m/%y', 'e': '%d-%m-%y', 'f': r'%d\%m\%y',
                   'g': '%d de %b de %Y', 'h': '%d de %B de %Y'
    }
    data = data.strip()
    data_traduzida = None
    if  match(r'(\d+/\d+/\d\d\d\d)',data):
        data_traduzida = padroes_data.get('a')
    elif match(r'(\d+/\d+/\d\d)$',data):
        data_traduzida = padroes_data.get('d')
    elif match(r'(\d+-\d+-\d\d\d\d)',data):
        data_traduzida = padroes_data.get('b')
    elif match(r'(\d+-\d+-\d\d)$',data):
        data_traduzida = padroes_data.get('e')
    elif match(r'(\d+\\\d+\\\d\d\d\d)',data):
        data_traduzida = padroes_data.get('c')
    elif match(r'(\d+\\\d+\\\d\d)$',data):
        data_traduzida = padroes_data.get('f')
    elif match(r'(\d+ de .+ de \d\d\d\d)',data):
        data_traduzida = padroes_data.get('h')
    elif match(r'(\d+ de .+ de \d\d)$',data):
        data_traduzida = padroes_data.get('g')

    if data_traduzida:
        return datetime.strptime(data, data_traduzida)
    else:
        return ''


def buscar_ou_inserir_obra(id_pessoa, id_processo, id_endereco, dados_obra, CadMunicipal):
    # Primeiro, tente encontrar uma obra existente com o mesmo protocolo
    #obra = db.Obras(db.Obras.Protocolo == id_processo)
    CadMunicipal = sub( '[^0-9]','', CadMunicipal)
    if  CadMunicipal.startswith('0'):
        CadMunicipal = CadMunicipal[1:]

    obra = db.executesql(f'''SELECT Id FROM Obras WHERE Obras.Protocolo = '{id_processo}' OR
                                Obras.protocolo_dof =  '{id_processo}' OR
                                Obras.protocolo_grcc =  '{id_processo}' OR
                                Obras.IdEndereco = '{id_endereco}' OR
                                Obras.CadMunicipal = '{CadMunicipal}' OR
                                Obras.CadMunicipal = '{"0"+CadMunicipal}' LIMIT 1 ;''')


    if obra:
        tipo = db.Processos(db.Processos.id == id_processo).IdTipo
        reg_obra = db.Obras(db.Obras.Id == obra[0][0])

        if tipo == 17 and reg_obra.protocolo_grcc is None:
            try:
                reg_obra.update_record(protocolo_grcc = id_processo)
                db.commit()
            except Exception as e:
                print('Erro ao atualizar protocolo da obra')
        if tipo == 3 and reg_obra.protocolo_dof is None:
            try:
                reg_obra.update_record(protocolo_dof = id_processo)
                db.commit()
            except Exception as e:
                print('Erro ao atualizar protocolo da obra')


        session.flash = f'Obra existente encontrada, id: {obra}'
        print(f'Obra ::{obra[0][0]}:: encontrada no banco de dados!')
        return obra[0][0]
    else:
        # Caso contrário, prepare os dados para inserção

        data_avara = dados_obra.get('data_alvara') or dados_obra.get('DATA DO ALVARÁ') or  dados_obra.get('DATA DO ALVARÁ (CONSTRUÇÃO/DEMOLIÇÃO)') or ''
        data_avara = traduz_data(data_avara)
        alvara = dados_obra.get('alvara')  or  dados_obra.get('ALVARÁ') or\
            dados_obra.get('ALVARÁ DE CONSTRUÇÃO') or dados_obra.get('ALVARÁ CONSTRUÇÃO E OU DEMOLIÇÃO') or ''

        if type(alvara) != tuple:
            alvara = alvara.split('/')[0]
        else:
            alvara = alvara[0]

        if 'MENTO(X)' in str(dados_obra.get('POSSUI DECK EM MADEIRA')).replace(' ','').upper():
            deck = 'Reflorestada'
        elif 'NATIVA(X)' in str(dados_obra.get('POSSUI DECK EM MADEIRA')).replace(' ','').upper():
            deck = 'Nativa'
        else:
            deck = ''
        if 'MENTO(X)' in str(dados_obra.get('POSSUI PERGOLADO EM MADEIRA')).replace(' ','').upper():
            pergolado = 'Reflorestada'
        elif 'NATIVA(X)' in str(dados_obra.get('POSSUI PERGOLADO EM MADEIRA')).replace(' ','').upper():
            pergolado = 'Nativa'
        else:
            pergolado = ''
        corte = dados_obra.get('VOLUME DE CORTE DE SOLO') or '0'
        aterro = dados_obra.get('VOLUME DE ATERRO DE SOLO') or '0'
        AreaTerreno = dados_obra.get('area_do_terreno') or \
                 dados_obra.get('ÁREA DO TERRENO') or '0'
        AreaConstrExist = dados_obra.get('area_existente') or \
                 dados_obra.get('ÁREA EXISTENTE') or dados_obra.get('ÁREA CONSTRUÍDA EXISTENTE') or '0'
        AreaConstrDemolir = dados_obra.get('area_demolir') or \
                 dados_obra.get('ÁREA A DEMOLIR') or dados_obra.get('ÁREA DEMOLIDA') or '0'
        AreaConstrExecutar = dados_obra.get('area_construida_final') or \
                dados_obra.get('ÁREA CONSTRUÍDA FINAL') or '0'

        dados_inserir = {
            'Protocolo': id_processo,
            'CadMunicipal': CadMunicipal, 'Alvara': alvara,
            'DataAlvara': data_avara ,
            'IdGerador': id_pessoa,
            'IdEndereco': id_endereco,
            'Finalidade': dados_obra.get('finalidade') or dados_obra.get('FINALIDADE') or '',
            'Nquartos': int(sub( '[^0-9]','',dados_obra.get('NÚMERO DE QUARTOS') or '0')) or None,
            'Piscina':  'on' if 'SIM(X)' in str(dados_obra.get('POSSUI PISCINA')).replace(' ','').upper() else None,
            'MadeiraReflorest' : 'on' if 'MENTO(X)' in str(dados_obra.get('POSSUI COBERTURA EM MADEIRA')).replace(' ','').upper() else None,
            'Pergolado' : pergolado,
            'Deck' : deck,
            'AreaTerreno': float( sub(',','.', sub('[^0-9,]','', AreaTerreno.split('m')[0]))),
            'AreaConstrExist': float( sub(',','.', sub('[^0-9,]','', AreaConstrExist.split('m')[0]))),
            'AreaConstrDemolir': float( sub(',','.', sub('[^0-9,]','', AreaConstrDemolir.split('m')[0]))),
            'AreaConstrExecutar': float( sub(',','.', sub('[^0-9,]','', AreaConstrExecutar.split('m')[0]))),
            'PavtosSubS': int(sub('[^0-9]','', dados_obra.get('pavimentos_sub') or
                dados_obra.get('PAVIMENTOS SUB') or '0')),
            'PavtosSobreS': int(sub('[^0-9]','', dados_obra.get('pavimentos_sup') or
                dados_obra.get('PAVIMENTOS') or '0')),
            'Corte': float( sub(',','.', sub('[^0-9,]','', corte.split('m')[0]))),
            'Aterro': float( sub(',','.', sub('[^0-9,]','', aterro.split('m')[0]))),
        }

        # Tente inserir a nova obra
        try:
            id_obra = db.Obras.validate_and_insert(**dados_inserir)
            print(f'Obra ::{id_obra}:: inserida no banco de dados!')
            db.commit()
            return id_obra.id
        except Exception as e:
            # Em caso de erro, registre o erro e redirecione para a página de obras
            print(f"Erro ao registrar obra: {e}")
            session.flash = f"Erro ao registrar obra: {e}"
            redirect(URL('default', 'Obras'))


def preparar_dados(dados_do_processo):
    caminho = Path(pasta_qualidade_habite_se) # type: ignore
    criar_pasta_compartilhada(dados_do_processo, caminho, list(pastas_aprova), numerar=True) # type: ignore
    id_pessoa = buscar_ou_inserir_pessoa(dados_do_processo['dados_gerador'])
    id_processo = buscar_ou_inserir_processo(dados_do_processo['protocolo'], id_pessoa, '31304', 3)
    id_logradouro = buscar_ou_inserir_logradouro(dados_do_processo['endereco_obra'])
    id_endereco = buscar_ou_inserir_endereco(id_logradouro, dados_do_processo['endereco_obra'])

    return (id_pessoa, id_processo, id_endereco)





@auth.requires_login()
def extrator_de_pdf(): #Menu
    # from myPDFextract import extract_text # type: ignore
    # import os

    # form_arquivo_processo = SQLFORM.factory(
    #     Field('Arq_mod', type='upload', uploadfolder=Path(request.folder, 'static/temp'),
    #           requires=IS_EMPTY_OR(IS_UPLOAD_FILENAME(extension='pdf')), label='Procurar Modelo'))

    # if form_arquivo_processo.process().accepted:
    #     arquivo_pdf = Path(request.folder, 'static/temp', form_arquivo_processo.vars.Arq_mod)
    #     texto = extract_text(arquivo_pdf)
    #     os.remove('{}/{}'.format(str(os.path.join(request.folder, 'static/temp')), form_arquivo_processo.vars.Arq_mod))
    #     for p,_ in enumerate(texto):
    #             texto[p] = texto[p].replace('SECRETARIA MUNICIPAL DE MEIO AMBIENTE E URBANISMO  Av. Lino José de Seixas, nº 861, Jd. Seixas CEP: 15061-060   Fone: (17) 3202-4010 qualidadeambiental@riopreto.sp.gov.br        PROJETO DE GERENCIAMENTO DE RESÍDUOS DA CONSTRUÇÃO CIVIL Informações Básicas Obrigatórias - § 3º do artigo 27, do Decreto nº 12.765, de 08 de abril de 2005  Lei Municipal nº 9393 de 20 de dezembro de 2004.', '')
    #             texto[p] = list(str(texto[p].split(': ')).split('m3 '))
    #     #TODO: Faça algo com o texto extraído

    # if 'texto' in locals():
    #     return dict(form=form_arquivo_processo,texto=LI(texto))
    # else:
    #     return dict(form=form_arquivo_processo)
    return dict(msg=mensagem_contrução)


@auth.requires_login()
def importar_processos_online(): #Menu # type: ignore
    from pathlib import Path
    from scrap_sigm_gest import pega_conteudo_html, captura_dados_de_processos # type: ignore
    import os
    import openpyxl # type: ignore
    from datetime import datetime as dt

    planilha_control = Path(pasta_qualidade, 'DOCUMENTOS', 'Controle de Processos.xlsx') # type: ignore
    wb = openpyxl.load_workbook(planilha_control)
    ws = wb.active
    hj = dt.today()
    plan_protocs = [r[0] for r in ws.iter_rows(5, ws.max_row, 2, 2, True)]

    form_arquivo = SQLFORM.factory(
        Field('Arq_mod', type='upload', required=True, uploadfolder=Path(request.folder, 'static/temp'),
              requires=IS_EMPTY_OR(IS_UPLOAD_FILENAME(extension='htm')), label='Anexar Pagina'),
        Field('Aba', type='string', requires=IS_IN_SET(['Pendentes', 'Meus', 'Todos']))
              )
    attributo_id_das_tables = {'Pendentes':'tablePendentes', 'Meus': 'tableRecebidos', 'Todos': 'tableTodos'}

    dict_processos_obtidos = {}
    if form_arquivo.process().accepted:
        table_selecionada = attributo_id_das_tables[form_arquivo.vars.Aba]
        arquivo_ = Path(request.folder, 'static/temp', form_arquivo.vars.Arq_mod)
        caminho_a = 'file:///{}'.format(arquivo_.absolute())
        conteudo = pega_conteudo_html(caminho_a)
        os.remove('{}/{}'.format(str(os.path.join(request.folder, 'static/temp')), form_arquivo.vars.Arq_mod  ))

        processos_obtidos = captura_dados_de_processos(conteudo, idtable=table_selecionada)
        string_nome_pgr = 'PROJETO DE GERENCIAMENTO DE RESÍDUOS DA CONSTRUÇÃO CIVIL'

        for i in processos_obtidos:
            if i['servico'] == string_nome_pgr or i['servico'] == "Projeto de Gerenciamento de Resíduos da Construção Civil":
                try:
                    l = ['X']
                    existe_pasta = criar_pasta_compartilhada(i, pasta_qualidade_pgr_ano, l , numerar=False) # type: ignore
                except Exception as e:
                    print(f"{e}")
                serv = 4
                depto = '1024412'
            elif i['servico'] in ('LICENÇA DE PUBLICIDADE', 'Licença de Publicidade') :
                try:
                    existe_pasta = criar_pasta_compartilhada(i,
                    'F:\\Qualidade Ambiental\\LICENÇA - PUBLICIDADES\\ANALISE DE PROTOCOLO (ONLINE)\\2024',  [''])
                    serv = 16
                    depto = '31304'
                except Exception as e:
                    print(f'Não foi possível criar pasta devido à: {e}.')

            elif i['servico'] in ('Análise do Documento de Origem Florestal', 'Análise do Documento de Origem Florestal'.upper()):
                serv = 3
                depto = '31304'
            elif i['servico'] in ('Análise de Gerenciamento de Resíduo de Construção', 'Análise de Gerenciamento de Resíduo de Construção'.upper()):
                serv = 17
                depto = '1024412'
            print('-'*20)
            print(i.get('dados_obs'))
            try:
                idpessoa = buscar_ou_inserir_pessoa(i)
            except Exception as e:
                print(f'Não foi possível buscar/inserir pessoa devido à: {e}')
            try:
                idproc = buscar_ou_inserir_processo(i.get('protocolo'), idpessoa, depto, serv)
            except Exception as e:
                print(f'Não foi possível buscar/inserir processo devido à: {e}')

            if serv == 3:
                try:
                    endereco_obra = i.get('dados_obs')
                except Exception as e:
                    print(f'Não foi possível obter o campo obs deste processo devido à: {e}')
                try:
                    idlogradouro = buscar_ou_inserir_logradouro(endereco_obra)
                except Exception as e:
                    print(f'Não foi possível buscar/inserir logradouro devido à: {e}')
                try:
                    idendereco = buscar_ou_inserir_endereco(idlogradouro, endereco_obra)
                except Exception as e:
                    print(f'Não foi possível buscar/inserir endereco devido à: {e}')
                cadmun = endereco_obra.get('CADASTRO MUNICIPAL') or endereco_obra.get('CADASTRO MUNICIPAL DO IMÓVEL')
                try:
                    idobra = buscar_ou_inserir_obra(idpessoa, idproc, idendereco, endereco_obra, cadmun)
                except Exception as e:
                    print(f'Não foi possível buscar/inserir obra devido à: {e}')
                print('-'*20)
            if serv == 17:
                try:
                    endereco_obra = i.get('dados_obs')
                except Exception as e:
                    print(f'Não foi possível obter o campo obs deste processo devido à: {e}')
                try:
                    idlogradouro = buscar_ou_inserir_logradouro(endereco_obra)
                except Exception as e:
                    print(f'Não foi possível buscar/inserir logradouro devido à: {e}')
                try:
                    idendereco = buscar_ou_inserir_endereco(idlogradouro, endereco_obra)
                except Exception as e:
                    print(f'Não foi possível buscar/inserir endereco devido à: {e}')

                cadmun = endereco_obra.get('CADASTRO MUNICIPAL') or endereco_obra.get('CADASTRO MUNICIPAL DO IMÓVEL')

                try:
                    idobra = buscar_ou_inserir_obra(idpessoa, idproc, idendereco, endereco_obra, cadmun)

                except Exception as e:
                    print(f'Não foi possível buscar/inserir obra devido à: {e}')

                print('-'*20)
                #redirect(URL('default', 'Obras', args=[idobra], vars={'f': 'ver'}))

            dict_processos_obtidos[A(i['protocolo'], _href=URL(c='default', f='Processos',
            args= idproc ,vars={'f':'editar'}))] = {'Serviço':i['servico'],'Nome': i['nome'], 'CPF': i['cpf'], 'Email': i['email']}
        try:
            for i in processos_obtidos:
                if i['protocolo'] not in plan_protocs:
                    ws.append([hj, i['protocolo'], i['nome'], 1 if serv == 3 else 0, 1 if serv == 17 else 0, 1 if serv == 4 else 0])
            wb.save(planilha_control)
        except Exception as e:
            print(f'{e}')

    try:
        session.dict_processos_obtidos = dict_processos_obtidos
        dict_processos_obtidos = session.dict_processos_obtidos
    except Exception as e:
        print(f'Erro: {e}')
        response.flash(f'{e}')

    if 'processos_obtidos' in locals():
        return response.render(form=form_arquivo,texto=BEAUTIFY(TABLE(*[TR(TD(k), TD(BEAUTIFY(TABLE(v) )) )  for (k, v) in dict_processos_obtidos.items()])))
    else:
        return dict(form=form_arquivo)


@auth.requires_login()
def Gerador_de_Docs(): #Menu
    '''Cria formulários automáticos para documentos com campos no formato : {campo}.
    Pede um documento de modelo .odt, neste modelo devem existir campos inseridos como variáveis (No libreoffice use ctrl+F2)
    Defina o valor do campo no documento no padrão entre colchetes: {campo}.
    Utilizando o módulo geradocspy.py importado de modelos,
    será gerado um formulário com os campos. Os valores poderão ser pré-preenchidos por registros no banco de dados selecionando o Protocolo'''

    import geradocspy # type: ignore
    from scrap_sigm_gest import pega_conteudo_html    # type: ignore  # Modulo gerador de documentos.odt
    from pathlib import Path            # Modulo de acesso ao sistema local de arquivos
    pasta = Path(request.folder)        # Para onde o arquivo vai
    import re
    import datetime
    my_regex = "( - )(\\d{3}\\.\\*{3}\\.\\*{3}\\-\\*{2})|( - )(\\d{2}\\.\\*{3}\\.\\*{3}\\/\\*{4}\\-\\*{2})"
    db.Obras.AreaTerreno.filter_in = lambda row: str(row).replace(',','.') if row else ''
    db.Obras.AreaTerreno.filter_out = lambda row: str(row).replace('.',',') if row else ''
    db.Obras.AreaConstrExist.filter_out = lambda row: str(row).replace('.',',') if row else ''
    db.Obras.AreaConstrDemolir.filter_out = lambda row: str(row).replace('.',',') if row else ''
    db.Obras.AreaConstrExecutar.filter_out = lambda row: str(row).replace('.',',') if row else ''
    db.Obras.Corte.filter_out = lambda row: str(row).replace('.',',') if row else ''
    db.Obras.Aterro.filter_out = lambda row: str(row).replace('.',',') if row else ''
    db.Pgrcc.cls_a.filter_out = lambda row: str(row).replace('.',',') if row else ''
    db.Pgrcc.cls_b.filter_out = lambda row: str(row).replace('.',',') if row else ''
    db.Pgrcc.cls_d.filter_out = lambda row: str(row).replace('.',',') if row else ''


    form_arquivo_processo = SQLFORM.factory(
        Field('ref_arq_modelo_doc', type='list:reference Modelos_de_docs',
        requires=IS_EMPTY_OR(IS_IN_DB(db, 'Modelos_de_docs.id', db.Modelos_de_docs._format)), label='Selecionar Modelo Salvo'),
        Field('Arq_mod', type='upload', uploadfolder=Path(pasta, 'static/temp'),
                requires = IS_EMPTY_OR([IS_UPLOAD_FILENAME(extension='[odt]|[html]')]), length= 30, tablename='n', filename= 'x',
                 label='ou Enviar Modelo'),
        Field('Protocolo', ),
        Field('Salvar_modelo', type='boolean', ) ,
        Field('Tipo_do_Processo', 'list:string', requires=IS_IN_DB(db, 'Servicos.id', db.Servicos._format)),
        upload=URL('download'),
        formname = 'FormDeArq', formstyle='bootstrap3_inline', ) # Formulário de docs e processos

    #TODO: Os campos do formulário deve vir na ordem em que são coletados do documento
    #TODO: O arquivo modelo enviado deve ser deletado de static/temp apos usado
    #TODO: O arquivo gerado deve ser aberto ou haver um link para abrir logo apos gerado.
    #TODO: O arquivo gerado deve ir para pasta certa ao ser gerado.

    arquivo_mod = None
    session.campos_form = campos_form = {}
    session.conteudo = None



    campos = {}
    if form_arquivo_processo.process(keepvalues = True).accepted:
        encontrado = db.Processos(db.Processos.Protocolo.contains(form_arquivo_processo.vars.Protocolo.strip()))
        session.protoc = encontrado.id
        session.Arq_mod = form_arquivo_processo.vars.Arq_mod
        if not encontrado:
            session.flash = f"Protocolo {form_arquivo_processo.vars.Protocolo} não encontrado no Banco de Dados."
            redirect(URL('Ferramentas', 'Gerador_de_Docs'))
        else:
            form_arquivo_processo.vars.Protocolo = encontrado.id
            if form_arquivo_processo.vars.Salvar_modelo == True and form_arquivo_processo.vars.Arq_mod!= None:
                nome_do_arq_modelo = request.vars.Arq_mod.filename
                arquivo_id = db.Modelos_de_docs.update_or_insert(
                db.Modelos_de_docs.nome_modelo_doc == nome_do_arq_modelo,
                arq_modelo_doc   = open('{}'.format( (Path(pasta, 'static/temp', session.Arq_mod).absolute()) ), 'w' ),
                nome_modelo_doc  = nome_do_arq_modelo,
                servico_refere   = form_arquivo_processo.vars.Tipo_do_Processo)
                session.arquivo_id = arquivo_id




        if  form_arquivo_processo.vars.ref_arq_modelo_doc != None:
            arquivo_nome = db.Modelos_de_docs(db.Modelos_de_docs.id == form_arquivo_processo.vars.ref_arq_modelo_doc).arq_modelo_doc
            arquivo_mod = Path(pasta,  pasta_de_modelos_do_servidor_de_arquivos,  Path(arquivo_nome)) # type: ignore
            session.Arq_mod = arquivo_mod

        else:
            arquivo_mod = Path(pasta,  'static/temp', Path(session.Arq_mod))
            session.Arq_mod = arquivo_mod

        try:
            if 'html' in arquivo_mod.suffix:
                conteudo = pega_conteudo_html('file:///{}'.format(arquivo_mod.absolute()))
                myregex = re.compile('\\{.*\\}')
                campos = conteudo.find_all(text=myregex)
                campos_do_doc = [c.replace('{', '').replace('}', '') for c in campos]
                session.conteudo = conteudo.decode()
                response.flash = f"campos: {campos}"
            else:
                with open('{}'.format(arquivo_mod), 'rb') as arquivo_mod:
                    conteudo = geradocspy.ler_o_arq_modelo(arquivo_mod)
                    session.conteudo = conteudo
                    campos_do_doc = geradocspy.pegar_campos_do_modelo(session.conteudo)
                    campos_do_doc = [c.replace('{', '').replace('}', '') for c in campos_do_doc]
                    session.campos_do_doc = campos_do_doc
                    response.flash = f"Enviando Arquivo de Modelo. {session.campos_do_doc}"

        except Exception as e:
            response.flash = "Erro ao enviar arquivo: Error: --> {}".format( e)

        if session.protoc:
            Gerador = db.Pessoas.with_alias('gerador')
            RespTec = db.Pessoas.with_alias('responsavel')
            # Publicidade
            if form_arquivo_processo.vars.Tipo_do_Processo == 16:
                query = db((db.Processos.id == session.protoc) & ( db.Processos.IdPessoa == db.Pessoas.Id) \
                & (db.Publicidades.Protocolo == db.Processos.id )).select()

            # PGRCC
            elif form_arquivo_processo.vars.Tipo_do_Processo in [4]:
                query = db((db.Processos.id == session.protoc) &\
                        (db.Processos.IdPessoa == db.Pessoas.Id) & (db.Obras.Protocolo == db.Processos.id) &\
                        (db.Pgrcc.protocolo == db.Processos.id)).select(db.Obras.Id,
                         db.Obras.Protocolo, db.Obras.CadMunicipal, db.Obras.Corte, db.Obras.Aterro , db.Obras.Alvara, db.Obras.DataAlvara, db.Obras.IdEndereco ,
                         db.Obras.Finalidade, db.Obras.AreaTerreno, db.Obras.AreaConstrExecutar, db.Obras.AreaConstrDemolir,  db.Obras.AreaConstrExist,
                         db.Obras.modified_on,  db.Pgrcc.ALL, Gerador.Nome ,
                         RespTec.celular, RespTec.Email, RespTec.RegistroProf, RespTec.Nome, left=(Gerador.on(Gerador.id==db
                    .Pgrcc.idgerador), RespTec.on(RespTec.id == db.Pgrcc.resptecnico)  )  )

             # Análise do GRCC
            elif form_arquivo_processo.vars.Tipo_do_Processo in [17]:
                query = db((db.Processos.id == session.protoc) &\
                        (db.Processos.IdPessoa == db.Pessoas.Id) & ((db.Obras.Protocolo == db.Processos.id) | (db.Obras.protocolo_grcc == db.Processos.id)) &\
                        (db.Pgrcc.idobra == db.Obras.id) & (db.Analise_GRCC.idobra == db.Obras.id)
                        ).select(db.Obras.Id,
                         db.Obras.Protocolo, db.Obras.CadMunicipal, db.Obras.Alvara, db.Obras.DataAlvara, db.Obras.IdEndereco ,
                         db.Obras.Finalidade, db.Obras.AreaTerreno, db.Obras.AreaConstrExecutar, db.Obras.Piscina,
                         db.Obras.CobertMetalica, db.Obras.MadeiraReflorest, db.Obras.Deck, db.Obras.Pergolado,
                         db.Obras.modified_on, db.Analise_GRCC.ALL,  db.Pgrcc.ALL, Gerador.Nome ,
                         RespTec.celular, RespTec.Email, RespTec.RegistroProf, RespTec.Nome, left=(Gerador.on(Gerador.id==db
                    .Pgrcc.idgerador), RespTec.on(RespTec.id == db.Pgrcc.resptecnico)))

            # DOF
            elif form_arquivo_processo.vars.Tipo_do_Processo  == 3:
                query = db((db.Processos.id == session.protoc) &\
                        (db.Processos.IdPessoa == db.Pessoas.Id) & ((db.Obras.Protocolo == db.Processos.id)
                        | (db.Obras.protocolo_dof == db.Processos.id)) ).select( db.Obras.Id,
                         db.Obras.Protocolo, db.Obras.CadMunicipal, db.Obras.Alvara, db.Obras.DataAlvara, db.Obras.IdEndereco ,
                         db.Obras.Finalidade, db.Obras.AreaTerreno, db.Obras.AreaConstrExecutar, db.Obras.Piscina,
                         db.Obras.CobertMetalica, db.Obras.MadeiraReflorest, db.Obras.Deck, db.Obras.Pergolado, db.Obras.modified_on,
                         Gerador.Nome, RespTec.celular, RespTec.Email, RespTec.RegistroProf, RespTec.Nome, left=(Gerador.on(Gerador.id==db
                    .Processos.IdPessoa), RespTec.on(RespTec.id == db.Obras.resptecnico)))
            else:
                response.flash = f"Não foi possível definir o tipo de processo do documento. {session.campos}"


        try:
            data = query.render(0).as_dict()
            data['responsavel']['Nome_resp'] = data['responsavel'].pop('Nome')

            if data:
                # Process data based on field types
                for tabela in data.values():

                    for field, value in tabela.items():
                        if field in campos_do_doc:
                            if isinstance(value, datetime.datetime) or isinstance(value, datetime.date):
                                value = value.strftime("%d/%m/%Y")
                            elif isinstance(value, (int, float)) and (value in [0, 0.0]):
                                value = '0,00'
                            value = '' if value is None else value
                            campos_form[field] = value

            for field in campos_do_doc:
                if field not in campos_form:
                    if field == 'Despacho':
                        from despachos import Despachar_dof  # type: ignore
                        query_obra = db.Obras(db.Obras.id == data.get('Obras').get('Id'))
                        texto_despacho = Despachar_dof(data, query_obra, imprime_cabacalho=False)
                        campos_form[field] = texto_despacho
                    elif field == 'Despacho_Agrcc':
                        from despachos import Despachar_Agrcc  # type: ignore
                        campos_form[field] = Despachar_Agrcc()
                    else:
                        campos_form[field] = '''\n'''

        except Exception as e:
            response.flash = "Erro ao atribuir dados do processo: Error: {} --> {}".format(e, campos_form)

        session.campos_form = campos_form

    elif form_arquivo_processo.errors:
        response.flash = "Verifique os erros no formulário"
    else:
        pass
    #del(session.Arq_mod, session.conteudo, session.valores_de_entrada, session.campos_form)
    return dict(form_arquivo_processo=form_arquivo_processo)



@auth.requires_login()
def Form_Gerador_de_Docs():
    import geradocspy # type: ignore

    campos_form = session.campos_form

    link_para_download = None
    formulario = None
    arquivo_final = None
    if campos_form:
        try:
            formulario = SQLFORM.dictform(campos_form ,
            formname='formulario', tablename='FormTable', )
        except Exception as e:
            response.flash(e)

    if formulario and formulario.process(keepvalues = True).accepted:
        valores_de_entrada = {}
        for k, v in formulario.vars.items( ):
            valores_de_entrada[k] = v
        session.valores_de_entrada = valores_de_entrada

        try:

            if 'html' in session.Arq_mod.suffix:
                #arquivo_final = geradocspy.preencher_conteudo(conteudo = session.conteudo, entradas_dos_campos = session.valores_de_entrada)
                arquivo_final = f'{session.conteudo.format(**session.valores_de_entrada)}'
                session.arquivo_final = arquivo_final

            else:
                arquivo_final = geradocspy.preencher_arquivo_novo(
                campo_p_nome = str(valores_de_entrada.get('Nome')).replace(' ', '_'),
                conteudo = session.conteudo,
                entradas_dos_campos = session.valores_de_entrada,
                arquivo_mod = session.Arq_mod)

            session.nome_arquivo_final = str(valores_de_entrada.get('Nome')).replace(' ', '_') + '.odt'
            response.flash = "Arquivo Preenchido e Salvo"

        except Exception as e:
            print(e)
            response.flash = f'Erro ___ {e}'

    elif formulario and formulario.errors:
        response.flash = "Verifique os erros no formulário"
    else:
        pass

    #TODO: O arquivo gerado deverá ser registrado em banco vinculado ao protocolo utilizado
    #TODO: Botão com link para acessar o arquivo gerado deve funcionar.


    #Caminho_pasta = LOAD(c='Ferramentas',f='Pastas_Pgrccs.load', args=['2023/0_RECEBIDOS/'], ajax=True )

    if 'html' in session.Arq_mod.suffix:
        return dict(formulario=formulario if 'formulario' in locals() else None, arquivo_final=arquivo_final)
    else:

        return dict(formulario=formulario if 'formulario' in locals() else None, arquivo_final=session.nome_arquivo_final, )


# @auth.requires_login()
# def Pastas_Pgrccs():
#     from gluon.tools import Expose

#     return dict(files=Expose('F:\\Qualidade Ambiental\\PGRCC - GESTÃO DIGITAL'))

# @auth.requires_login()
# def Pastas_Publicidades():
#     from gluon.tools import Expose
#     return dict(files=Expose('F:\\Qualidade Ambiental\\LICENÇA - PUBLICIDADES'))




def Importar_do_Aprova(): #Menu
    formarqhtml = SQLFORM.factory(Field('arquivo',
                                type ='upload', label = "Inserir a página html salva em seu computador."),
                                formstyle = 'table3cols')
    dados_do_processo = None

    if formarqhtml.process().accepted:
        #session.flash = 'Enviando, aguarde'
        arquivo = formarqhtml.vars.arquivo
        dados_do_processo = processar_arquivo_html(arquivo) or None
        if dados_do_processo:
            remover_arquivo(arquivo)

            id_pessoa, id_processo, id_endereco = preparar_dados(dados_do_processo)
            id_obra = buscar_ou_inserir_obra(id_pessoa, id_processo, id_endereco,
                                             dados_do_processo['dados_obra'],
                                             dados_do_processo['endereco_obra']['CadMunicipal'])

            atualizar_planilha_control(dados_do_processo)

            redirect(URL('default', 'Obras',  args=[id_obra], vars={'f':'editar'}))
        else:
            session.flash = 'Erro ao processar o arquivo.'

    if dados_do_processo:
        return dict(dados_do_processo=dados_do_processo, formarqhtml=formarqhtml)
    else:
        return dict(formarqhtml=formarqhtml)





def Gerador_de_QrCode(): #Menu
    #import qrcode as qr
    #from PIL import Image
    from QRcodeGen import adjust_logo, add_qrdata, save_qrcode  # type: ignore
    from pathlib import Path


    form = SQLFORM.factory(
    Field('conteudo', type='text',label='texto ou link',formstyle='bootstrap3_inline' ),
    Field('Nome'),
    Field('logo', type='upload')
    )

    #codeqr = qr.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
    qrcode = session.qrcode or ''

    if form.process(keepvalues=True).accepted:
        logo = None
        session.qrcode = None
        if form.vars.logo:
            path_logo = open('{}'.format( (Path(request.folder, 'uploads', Path(f'{form.vars.logo}')  ) )), 'rb' )
            logo = adjust_logo(path_logo)

        response.flash =  f'enviado'
        qrcode = add_qrdata(f'{form.vars.conteudo}', 'Black', logo if logo else None)
        arquivo = Path(request.folder, 'static/qrcodes', f'{form.vars.Nome}.png')
        save_qrcode(qrcode, arquivo)
        img = IMG(_src=URL('static', 'qrcodes', args=f'{form.vars.Nome}.png'), _alt='img')
        session.qrcode = img

    return response.render(form=form, qrcode = session.qrcode, logo=logo if 'logo' in locals() else '')



def Despachar_Processos(): #Menu
    from gluon.contrib.markdown.markdown2 import MarkdownWithExtras as Markdown2
    from despachos import Despachar_dof  # type: ignore

    processo = int(request.vars.processo)
    query = db(db.Obras.Protocolo == processo).select().render(0).as_dict()
    query_obra = db.Obras(db.Obras.id == query.get('Id'))
    texto_despacho = Despachar_dof(query, query_obra)

    markdowner = Markdown2(html4tags=True, tab_width=4, )
    texto_md = markdowner.convert(texto_despacho) or None


    return dict(conteudo = XML(texto_md))